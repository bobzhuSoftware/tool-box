"""Excel string search — find strings across selected sheets of Excel files."""
import io
import os

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel

from app.core.auth import require_user
from app.core.db import User

router = APIRouter()


class ExcelStructureRequest(BaseModel):
    path: str


class ExcelSearchSheet(BaseModel):
    name: str  # sheet name
    columns: list[str] = []  # column letters to search; empty = all columns


class ExcelSearchTarget(BaseModel):
    file: str  # absolute path to an .xlsx/.xlsm file
    sheets: list[ExcelSearchSheet]  # sheets (with optional column filter) to search


class ExcelSearchRequest(BaseModel):
    queries: list[str] = []  # one or more strings to search for
    exact: bool = False  # True = whole-cell equals; False = case-insensitive substring
    targets: list[ExcelSearchTarget] = []


_EXCEL_SEARCH_MAX_MATCHES = 5000


@router.post("/api/excel-search/structure")
def excel_search_structure(req: ExcelStructureRequest, user: User = Depends(require_user)):
    """Scan a folder (recursively) and return each Excel file's sheets and columns.

    Each sheet includes its columns (column letter + first-row header text) so the
    frontend can let the user pick which sheets and which columns to search.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    root = req.path.strip().strip('"').strip("'")
    if not root:
        raise HTTPException(status_code=400, detail="路径不能为空")
    if not os.path.exists(root):
        raise HTTPException(status_code=400, detail=f"路径不存在: {root}")
    if not os.path.isdir(root):
        raise HTTPException(status_code=400, detail=f"不是文件夹: {root}")

    files: list[dict] = []
    errors: list[dict] = []

    for dirpath, _dirnames, filenames in os.walk(root):
        for fn in sorted(filenames):
            if fn.startswith("~$"):
                continue  # skip Excel lock/temp files
            if not fn.lower().endswith((".xlsx", ".xlsm")):
                continue
            fpath = os.path.join(dirpath, fn)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
            except Exception as exc:  # noqa: BLE001 - report unreadable files, keep going
                errors.append({"file": fpath, "error": str(exc)})
                continue
            try:
                sheets_info = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    columns = []
                    for header_row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                        for idx, val in enumerate(header_row, start=1):
                            columns.append({
                                "letter": get_column_letter(idx),
                                "header": "" if val is None else str(val)[:60],
                            })
                        break  # only the first row
                    sheets_info.append({"name": sheet_name, "columns": columns})
            finally:
                wb.close()
            files.append({
                "file": fpath,
                "fileName": fn,
                "relDir": os.path.relpath(dirpath, root),
                "sheets": sheets_info,
            })

    return {
        "path": root,
        "fileCount": len(files),
        "files": files,
        "errors": errors,
    }


@router.post("/api/excel-search")
def excel_search(req: ExcelSearchRequest, user: User = Depends(require_user)):
    """Search one or more strings within the user-selected sheets of the given files.

    Returns each hit's file name, sheet name, cell address (e.g. H5) and which
    query matched, plus a per-query summary (including strings that were not found).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string

    # normalize + dedupe queries, keep original order
    queries: list[str] = []
    seen: set[str] = set()
    for q in req.queries:
        q = (q or "").strip()
        if q and q not in seen:
            seen.add(q)
            queries.append(q)
    if not queries:
        raise HTTPException(status_code=400, detail="搜索字符串不能为空")
    if not req.targets:
        raise HTTPException(status_code=400, detail="请先选择要搜索的 sheet")

    # precompute (original, lowercased) pairs for fast matching
    query_pairs = [(q, q.lower()) for q in queries]
    counts: dict[str, int] = {q: 0 for q in queries}

    matches: list[dict] = []
    errors: list[dict] = []
    files_scanned = 0
    sheets_scanned = 0
    truncated = False

    for target in req.targets:
        fpath = target.file
        wanted = [s for s in target.sheets if s and s.name]
        if not wanted:
            continue
        if not os.path.isfile(fpath):
            errors.append({"file": fpath, "error": "文件不存在"})
            continue
        fn = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - report unreadable files, keep going
            errors.append({"file": fpath, "error": str(exc)})
            continue
        files_scanned += 1
        try:
            available = set(wb.sheetnames)
            for sheet in wanted:
                sheet_name = sheet.name
                if sheet_name not in available:
                    errors.append({"file": fpath, "error": f"sheet 不存在: {sheet_name}"})
                    continue
                # build allowed column-index set (empty/None means all columns)
                allowed_cols = None
                if sheet.columns:
                    allowed_cols = set()
                    for letter in sheet.columns:
                        try:
                            allowed_cols.add(column_index_from_string(letter))
                        except (ValueError, TypeError):
                            pass
                ws = wb[sheet_name]
                sheets_scanned += 1
                for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    for cidx, val in enumerate(row, start=1):
                        if val is None:
                            continue
                        if allowed_cols is not None and cidx not in allowed_cols:
                            continue
                        sval = str(val)
                        sval_lower = sval.lower()
                        for q, q_lower in query_pairs:
                            hit = (sval == q) if req.exact else (q_lower in sval_lower)
                            if not hit:
                                continue
                            counts[q] += 1
                            matches.append({
                                "file": fpath,
                                "fileName": fn,
                                "sheet": sheet_name,
                                "cell": f"{get_column_letter(cidx)}{ridx}",
                                "value": sval[:200],
                                "query": q,
                            })
                            if len(matches) >= _EXCEL_SEARCH_MAX_MATCHES:
                                truncated = True
                                break
                        if truncated:
                            break
                    if truncated:
                        break
                if truncated:
                    break
        finally:
            wb.close()
        if truncated:
            break

    summary = [{"query": q, "count": counts[q]} for q in queries]
    not_found = [q for q in queries if counts[q] == 0]

    return {
        "queries": queries,
        "exact": req.exact,
        "filesScanned": files_scanned,
        "sheetsScanned": sheets_scanned,
        "matchCount": len(matches),
        "truncated": truncated,
        "matches": matches,
        "summary": summary,
        "notFound": not_found,
        "errors": errors,
    }


class ExcelExportMatch(BaseModel):
    fileName: str = ""
    sheet: str = ""
    cell: str = ""
    query: str = ""
    value: str = ""


class ExcelExportSummaryRow(BaseModel):
    query: str = ""
    files: str = ""


class ExcelExportRequest(BaseModel):
    matches: list[ExcelExportMatch] = []
    summary: list[ExcelExportSummaryRow] = []


@router.post("/api/excel-search/export")
def excel_search_export(req: ExcelExportRequest, user: User = Depends(require_user)):
    """Export search results as a single .xlsx workbook with two sheets:
    「明细」(every hit) and 「汇总」(per-string file list)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws_detail = wb.active
    ws_detail.title = "明细"
    ws_detail.append(["字符串", "文件", "Sheet", "单元格", "内容"])
    for cell in ws_detail[1]:
        cell.font = bold
    for m in req.matches:
        ws_detail.append([m.query, m.fileName, m.sheet, m.cell, m.value])

    ws_summary = wb.create_sheet("汇总")
    ws_summary.append(["字符串", "出现的文件"])
    for cell in ws_summary[1]:
        cell.font = bold
    for r in req.summary:
        ws_summary.append([r.query, r.files])

    # widen columns a bit for readability
    for ws, widths in ((ws_detail, [16, 28, 18, 10, 50]), (ws_summary, [16, 70])):
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=excel-search-result.xlsx"},
    )
